"""
tracker/__init__.py — logs every application to a local Excel file + sends daily digest email

LOGGING: Writes to logs/applications.xlsx by default — no Google Cloud setup
needed. This file lives in your repo and gets committed back by the GitHub
Actions workflow alongside the other state files, so it persists across runs
and you can open it directly in Excel/Numbers/Google Sheets import any time.

Google Sheets logging is still supported as an OPTIONAL alternative if
GSHEET_CREDENTIALS_JSON is set — see _log_to_sheets() — but it is no longer
required and is not used unless that secret is present.
"""
import os
from pathlib import Path
from datetime import datetime
from utils import log, get_config

ROOT       = Path(__file__).parent.parent
XLSX_FILE  = ROOT / "logs" / "applications.xlsx"

SHEET_HEADERS = [
    "Date", "Company", "Role", "Location", "Source",
    "Fit Score", "H1B Status", "H1B Category",
    "ATS Keywords", "URL", "PDF Path", "Status", "Notes"
]


def _get_or_create_workbook():
    """
    Opens logs/applications.xlsx, creating it with headers if it doesn't
    exist yet. Returns (workbook, worksheet).
    """
    from openpyxl import Workbook, load_workbook

    XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)

    if XLSX_FILE.exists():
        wb = load_workbook(XLSX_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(SHEET_HEADERS)
        # Bold header row + reasonable column widths for readability
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
        widths = [16, 28, 28, 18, 12, 9, 24, 16, 30, 40, 30, 10, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    return wb, ws


def _row_for(job: dict, pdf_path: str | None, submitted: bool) -> list:
    from h1b_lookup import lookup_h1b_history, get_h1b_label
    status       = "Applied" if submitted else "Dry Run"
    h1b_full     = job.get("h1b_history") or lookup_h1b_history(job.get("company", ""))
    h1b_category = get_h1b_label(job.get("company", ""))
    return [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        job.get("company", ""),
        job.get("title", ""),
        job.get("location", ""),
        job.get("source", ""),
        job.get("fit_score", ""),
        h1b_full,
        h1b_category,
        ", ".join(job.get("ats_keywords", [])),
        job.get("url", ""),
        pdf_path or "",
        status,
        job.get("fit_reasoning", ""),
    ]


def log_application(job: dict, pdf_path: str | None, submitted: bool):
    """
    Appends one row for this job to logs/applications.xlsx.
    This is the default, no-setup-required tracking method.
    """
    try:
        wb, ws = _get_or_create_workbook()
        ws.append(_row_for(job, pdf_path, submitted))
        wb.save(XLSX_FILE)
        h1b_full = job.get("h1b_history", "")
        log.info(f"  Logged: {job['title']} @ {job['company']} | {h1b_full}")
    except ImportError:
        log.warning(
            "Excel logging skipped: openpyxl not installed. "
            "Run: pip install openpyxl --break-system-packages"
        )
    except Exception as e:
        log.warning(f"Excel logging failed: {e}")

    # Optional: also log to Google Sheets if the user has set this up.
    # Entirely opt-in — skipped silently if the secret isn't present.
    if os.environ.get("GSHEET_CREDENTIALS_JSON"):
        _log_to_sheets(job, pdf_path, submitted)


def _log_to_sheets(job: dict, pdf_path: str | None, submitted: bool):
    """Optional secondary logging to Google Sheets, only runs if configured."""
    try:
        import gspread, json
        from google.oauth2.service_account import Credentials

        creds_dict = json.loads(os.environ["GSHEET_CREDENTIALS_JSON"])
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/spreadsheets",
                    "https://www.googleapis.com/auth/drive"]
        )
        gc = gspread.authorize(creds)
        cfg = get_config()
        sheet_name = cfg["tracking"]["sheet_name"]
        try:
            sh = gc.open(sheet_name).sheet1
        except Exception:
            sh = gc.create(sheet_name).sheet1
            sh.append_row(SHEET_HEADERS)
        sh.append_row(_row_for(job, pdf_path, submitted))
    except Exception as e:
        log.debug(f"Optional Google Sheets logging failed (non-fatal): {e}")


def send_run_notification(results: list[dict]):
    """
    Sends a concise email after each run using Gmail App Password (SMTP).

    Setup — 2 minutes, no Google Cloud needed:
      1. Go to myaccount.google.com/security
      2. Enable 2-Step Verification (if not already on)
      3. Go to myaccount.google.com/apppasswords
      4. App name: "ApplyRyt" → Create
      5. Copy the 16-character password shown
      6. Add as GitHub secret: GMAIL_APP_PASSWORD

    That's it. One secret, works immediately.
    """
    import smtplib
    from email.mime.text import MIMEText

    cfg     = get_config()
    dry_run = cfg.get("dry_run", True)
    applied = [r for r in results if r.get("submitted")]
    skipped = [r for r in results if not r.get("submitted")]

    if not applied:
        log.info("  No applications this run — skipping email notification")
        return

    # Get credentials — App Password only (no OAuth complexity)
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not app_password:
        log.warning(
            "Email skipped: GMAIL_APP_PASSWORD not set. "
            "Get one at myaccount.google.com/apppasswords → add as GitHub secret."
        )
        return

    sender = cfg["profile"]["email"]
    recipient = cfg["tracking"]["notify_email"]
    now_str = datetime.now().strftime("%b %d, %Y %I:%M %p UTC")
    action  = "Dry-run logged" if dry_run else "Applied"

    # H1B breakdown
    from collections import Counter, defaultdict
    h1b_counts = Counter(j.get("h1b_history","❓ No H1B Record") for j in applied)

    priority_jobs = [
        j for j in applied
        if "Exempt" in j.get("h1b_history","") or "Sponsor" in j.get("h1b_history","")
    ]
    priority_ids = {id(j) for j in priority_jobs}
    norecord_jobs = [j for j in applied if id(j) not in priority_ids]

    # Group priority jobs by the configured role they were searched under
    # (e.g. "Business Analyst", "Product Owner") so similar titles bucket together
    by_role = defaultdict(list)
    for j in priority_jobs:
        role_key = j.get("searched_role") or j.get("title", "Other")
        by_role[role_key].append(j)

    # Build email body
    lines = [
        f"{'[DRY RUN] ' if dry_run else ''}ApplyRyt Run Summary — {now_str}",
        "=" * 55,
        "",
        f"{action}: {len(applied)} jobs  |  Skipped: {len(skipped)}",
        "",
        "H1B SPONSORSHIP BREAKDOWN:",
        f"  🎓 Cap-Exempt        {h1b_counts.get('🎓 H1B Exempt (Cap-Free)', 0)} jobs  ← applied FIRST",
        f"  ✅ H1B Sponsor       {h1b_counts.get('✅ H1B Sponsor', 0)} jobs  ← applied SECOND",
        f"  ❓ No Record         {len(norecord_jobs)} jobs  (applied in background, not listed below)",
        "",
        "PRIORITY COMPANIES — Cap-Exempt + H1B Sponsor (grouped by role):",
        "=" * 55,
    ]

    for role_key in sorted(by_role.keys()):
        role_jobs = sorted(
            by_role[role_key],
            key=lambda x: (0 if "Exempt" in x.get("h1b_history","") else 1, -x.get("fit_score", 0))
        )
        lines.append("")
        lines.append(f"▶ {role_key}  ({len(role_jobs)} jobs)")
        lines.append("-" * 55)
        for j in role_jobs:
            h1b = j.get("h1b_history", "")
            tag = "🎓" if "Exempt" in h1b else "✅"
            company = j.get("company", "Unknown Company")
            lines.append(f"  {tag} {company}")
            lines.append(f"     Fit: {j.get('fit_score','?')}/10  |  {h1b}  |  {j.get('source','')}")
            lines.append(f"     {j.get('url', '')}")

    if norecord_jobs:
        lines.append("")
        lines.append(f"❓ {len(norecord_jobs)} No-Record jobs also applied this run (full list in logs/applications.xlsx)")

    lines += [
        "",
        "-" * 55,
        "View all: logs/applications.xlsx (in your repo, updated every run)",
        "",
        "Legend:",
        "  🎓 Cap-Exempt   = university/hospital/nonprofit — NO LOTTERY, applied first",
        "  ✅ H1B Sponsor  = proven H1B filer — applied second, April lottery applies",
        "  ❓ No Record    = not in USCIS data, still applied, ask in interview",
        "",
        "To pause: set dry_run: true in config.yaml → push to GitHub",
    ]

    subject = (
        f"{'[DRY RUN] ' if dry_run else ''}ApplyRyt: {len(applied)} jobs "
        f"{'logged' if dry_run else 'applied'} — {datetime.now().strftime('%b %d %I:%M%p')}"
    )

    try:
        msg = MIMEText("\n".join(lines), "plain")
        msg["From"]    = sender
        msg["To"]      = recipient
        msg["Subject"] = subject

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, app_password)
            server.sendmail(sender, recipient, msg.as_string())

        log.info(f"Run notification sent to {recipient}")

    except smtplib.SMTPAuthenticationError:
        log.warning(
            "Email failed: Gmail authentication error. "
            "Make sure GMAIL_APP_PASSWORD is correct and 2-Step Verification is enabled. "
            "Get a new App Password at myaccount.google.com/apppasswords"
        )
    except Exception as e:
        log.warning(f"Email notification failed: {e}")


# Alias for backward compatibility
send_daily_digest = send_run_notification
